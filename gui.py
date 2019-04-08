from tkinter import *
from tkinter import messagebox
from tkinter import ttk
from openpyxl import load_workbook
from random import randint


# ======================Back-end======================


def submit_form():   # called from def create account
    file_name = 'banks_database.xlsx'
    data = load_workbook(file_name)
    active_sheet = data.active

    row = 2
    column = 1

    while True:
        if active_sheet.cell(row=row, column=column).value:
            row += 1
            continue
        else:
            active_sheet.cell(row=row, column=column).value = entry_display_acc_no.get()
            column += 1
            active_sheet.cell(row=row, column=column).value = entry_fname.get()
            column += 1
            active_sheet.cell(row=row, column=column).value = entry_lname.get()
            column += 1
            active_sheet.cell(row=row, column=column).value = entry_init_deposit.get()
            column += 1

            if column == 5:
                break

    data.save(filename=file_name)
    messagebox.showinfo('Create Account', 'Account Created successfully!')
    create_account_window.destroy()


def generate_account_no():   # called from def create account
    entry_display_acc_no.delete(0, END)
    entry_display_acc_no.insert(END, randint(10000, 99999))


def deposit_funds():  # called from def authenticate deposit
    dict_depo = []
    row = 2
    column = 3

    while True:
        if sheet.cell(row=row, column=column).value == entry_second_nm.get():
            column += 1
            dict_depo.append(sheet.cell(row=row, column=column).value)
            dict_depo.append(entry_deposit.get())

            current = int(dict_depo[0]) + int(dict_depo[1])

            sheet.cell(row=row, column=column).value = current

            messagebox.showinfo('Deposit', 'Deposit successfully! New balance is: {}'.
                                format(sheet.cell(row=row, column=column).value))
            book.save(filename=file)
            deposit_cash_window.destroy()
            break
        else:
            if row != max_row:
                row += 1
                continue
            else:
                messagebox.showinfo('Not Available', 'This account does not exist!')
                deposit_cash_window.destroy()
                break


def withdraw_funds():  # called from def authenticate withdraw
    # ent_last_name, entry_withdraw, withdraw_cash_window

    dict_withdraw = []
    row = 2
    column = 3

    while True:
        if w_sheet.cell(row=row, column=column).value == ent_last_name.get():
            column += 1
            dict_withdraw.append(w_sheet.cell(row=row, column=column).value)
            dict_withdraw.append(entry_withdraw.get())

            if int(dict_withdraw[1]) < int(dict_withdraw[0]):
                current = int(dict_withdraw[0]) - int(dict_withdraw[1])
                w_sheet.cell(row=row, column=column).value = current
                messagebox.showinfo('Deposit', 'Withdrawal successfully! New balance is: {}'.
                                    format(w_sheet.cell(row=row, column=column).value))
            else:
                messagebox.showerror('Failed', 'Sorry, you have insufficient funds!')

            w_book.save(filename=file_nm)
            withdraw_cash_window.destroy()
            break
        else:
            if row != max_row:
                row += 1
                continue
            else:
                messagebox.showinfo('Not Available', 'This account does not exist!')
                deposit_cash_window.destroy()
                break


def authenticate_deposit():  # called from def deposit cash

    global sheet, book, file
    file = 'banks_database.xlsx'
    book = load_workbook(file)
    sheet = book.active

    row = 2
    column = 1

    while True:
        if sheet.cell(row=row, column=column).value == entry_acc_no.get():
            column += 1
            if sheet.cell(row=row, column=column).value == entry_first_nm.get():
                column += 1
                if sheet.cell(row=row, column=column).value == entry_second_nm.get():
                    # part of subsystem deposit cash
                    global frame_deposit
                    frame_deposit = ttk.LabelFrame(deposit_cash_window, text='Deposit Cash', width=450, height=180)
                    frame_deposit.place(x=25, y=200)

                    label_deposit = Label(frame_deposit, text='Amount to Deposit: ')
                    label_deposit.place(x=25, y=25)

                    global entry_deposit
                    entry_deposit_var = IntVar()
                    entry_deposit = Entry(frame_deposit, width=20, textvariable=entry_deposit_var)
                    entry_deposit.place(x=200, y=25)

                    button_deposit = Button(frame_deposit, text='Deposit Cash', width=25, height=2, bg='Gray',
                                            command=lambda: deposit_funds())
                    button_deposit.place(x=70, y=80)
                    break
        else:
            global max_row
            max_row = sheet.max_row

            if row != max_row:
                row += 1
                continue
            else:
                messagebox.showinfo('Not Available', 'This account does not exist!')
                deposit_cash_window.destroy()
                break


def authenticate_withdraw():   # called from def withdraw cash
    # ent_first_name, ent_last_name, ent_acc_no, withdraw_cash_window

    global file_nm, w_book, w_sheet
    file_nm = 'banks_database.xlsx'
    w_book = load_workbook(file_nm)
    w_sheet = w_book.active
    row = 2
    column = 1

    while True:
        if w_sheet.cell(row=row, column=column).value == ent_acc_no.get():
            column += 1
            if w_sheet.cell(row=row, column=column).value == ent_first_name.get():
                column += 1
                if w_sheet.cell(row=row, column=column).value == ent_last_name.get():
                    frame_withdraw = ttk.LabelFrame(withdraw_cash_window, text='Withdraw Cash', width=450, height=180)
                    frame_withdraw.place(x=25, y=200)

                    label_withdraw = Label(frame_withdraw, text='Amount to withdraw: ')
                    label_withdraw.place(x=25, y=25)

                    global entry_withdraw
                    entry_withdraw_var = IntVar()
                    entry_withdraw = Entry(frame_withdraw, width=20, textvariable=entry_withdraw_var)
                    entry_withdraw.place(x=200, y=25)

                    button_withdraw = Button(frame_withdraw, text='Withdraw Cash', width=25, height=2, bg='Gray',
                                             command=lambda: withdraw_funds())
                    button_withdraw.place(x=70, y=80)
                    break
        else:
            global max_row
            max_row = w_sheet.max_row

            if row != max_row:
                row += 1
                continue
            else:
                messagebox.showinfo('Not Available', 'This account does not exist!')
                withdraw_cash_window.destroy()
                break


def current_balance():  # called from def check_balance()
    #  check_balance_window, e_acc_no, w_sheet
    fl_nm = 'banks_database.xlsx'
    wk_book = load_workbook(fl_nm)
    wk_sheet = wk_book.active

    row = 2
    column = 1

    while True:
        if wk_sheet.cell(row=row, column=column).value == e_acc_no.get():
            column += 3
            messagebox.showinfo('Balance', 'Your current balance is: {}'.
                                format(wk_sheet.cell(row=row, column=column).value))
            check_balance_window.destroy()
            break
        else:
            maxi_row = wk_sheet.max_row
            if maxi_row != row:
                row += 1
                continue
            else:
                messagebox.showinfo('Balance', 'Your account number does not exist in our bank!')
                check_balance_window.destroy()


def exit_program():
    exit()


def about_us():
    about_window = Tk()
    about_window.title('About Pythonista bank')
    about_window.geometry('500x500')
    about_window.resizable(width=False, height=False)
    about_window.configure(bg='Gray')

    # frame
    frame_about = ttk.LabelFrame(about_window, text='About Us', width=400, height=400)
    frame_about.place(x=70, y=25)

    # scrollbar
    sb = Scrollbar(frame_about)
    sb.pack(side=RIGHT, fill=Y)

    lb = Listbox(frame_about, yscrollcommand=sb.set, bd=3, width=30, height=20)
    lb.pack()

    sb.configure(command=lb.yview)

    # about us information on the listbox
    lb.configure(bg='Light Blue')
    lb.insert(END, 'Welcome to Pythonista Bank of Kenya.')
    lb.insert(END, 'We are registered as a non-operating')
    lb.insert(END, 'holding company which started operations')
    lb.insert(END, 'as a licensed banking institution with effect')
    lb.insert(END, 'from January 1, 2016.')
    lb.insert(END, '')
    lb.insert(END, '')
    lb.insert(END, '')
    lb.insert(END, 'The holding company was set up')
    lb.insert(END, 'to among other things to enhance')
    lb.insert(END, 'the Group’s capacity to access')
    lb.insert(END, 'unrestricted capital and also ')
    lb.insert(END, 'enable investment in new ventures ')
    lb.insert(END, 'outside banking regulations.')

    about_window.mainloop()


def create_account():  # this function creates subsystem(create account)
    global create_account_window

    create_account_window = Tk()
    create_account_window.title('Create Account.')
    create_account_window.configure(bg='Gray')
    create_account_window.geometry('500x400')
    create_account_window.resizable(width=False, height=False)

    # FRAMES
    create_acc_frame = ttk.LabelFrame(create_account_window, text='Create Account', width=450, height=180)
    create_acc_frame.place(x=25, y=15)

    submit_form_frame = ttk.LabelFrame(create_account_window, text='Submit Form.', width=450, height=180)
    submit_form_frame.place(x=25, y=200)

    # label_and_entries
    label_fname = Label(create_acc_frame, text='First Name: ')
    label_fname.place(x=5, y=10)

    global entry_fname
    entry_fname_var = StringVar()
    entry_fname = Entry(create_acc_frame, width=20, textvariable=entry_fname_var)
    entry_fname.place(x=150, y=10)

    label_lname = Label(create_acc_frame, text='Last Name: ')
    label_lname.place(x=5, y=40)

    global entry_lname
    entry_lname_var = StringVar()
    entry_lname = Entry(create_acc_frame, width=20, textvariable=entry_lname_var)
    entry_lname.place(x=150, y=40)

    label_display_acc_no = Label(create_acc_frame, text='Account No: ')
    label_display_acc_no.place(x=5, y=70)

    global entry_display_acc_no
    entry_display_acc_no = Entry(create_acc_frame, width=20)
    entry_display_acc_no.place(x=150, y=70)

    label_init_deposit = Label(create_acc_frame, text='Init Deposit: ')
    label_init_deposit.place(x=5, y=100)

    global entry_init_deposit
    entry_init_deposit_var = IntVar()
    entry_init_deposit = Entry(create_acc_frame, width=20, textvariable=entry_init_deposit_var)
    entry_init_deposit.place(x=150, y=100)

    # button
    btn_generate_acc_number = Button(submit_form_frame, text='Generate Account\nNumber', width=15, height=5, bg='Gray',
                                     command=lambda: generate_account_no())
    btn_generate_acc_number.place(x=5, y=20)

    btn_submit_form = Button(submit_form_frame, text='Submit the\nForm', width=15, height=5, bg='Gray',
                             command=lambda: submit_form())
    btn_submit_form.place(x=250, y=20)

    create_account_window.mainloop()


def deposit_cash():  # this function creates subsystem(deposit cash)
    global deposit_cash_window
    deposit_cash_window = Tk()
    deposit_cash_window.title('Deposit Cash.')
    deposit_cash_window.configure(bg='Gray')
    deposit_cash_window.geometry('500x400')

    # FRAMES
    global frame_authentication
    frame_authentication = ttk.LabelFrame(deposit_cash_window, text='Authentication', width=450, height=180)
    frame_authentication.place(x=25, y=15)

    # labels, entries, and a button
    label_first_nm = Label(frame_authentication, text='First Name: ')
    label_first_nm.place(x=5, y=10)

    global entry_first_nm
    entry_first_nm_var = StringVar()
    entry_first_nm = Entry(frame_authentication, textvariable=entry_first_nm_var, width=20)
    entry_first_nm.place(x=150, y=10)

    label_last_nm = Label(frame_authentication, text='Last Name: ')
    label_last_nm.place(x=5, y=40)

    global entry_second_nm
    entry_second_nm_var = StringVar()
    entry_second_nm = Entry(frame_authentication, textvariable=entry_second_nm_var, width=20)
    entry_second_nm.place(x=150, y=40)

    label_acc_no = Label(frame_authentication, text='Acc Number: ')
    label_acc_no.place(x=5, y=70)

    global entry_acc_no
    entry_acc_no_var = StringVar()
    entry_acc_no = Entry(frame_authentication, textvariable=entry_acc_no_var, width=20)
    entry_acc_no.place(x=150, y=70)

    # Button for authenticating the user's account.
    btn_authenticate = Button(frame_authentication, text='Authenticate', width=30, bg='Gray',
                              command=lambda: authenticate_deposit())
    btn_authenticate.place(x=45, y=110)

    deposit_cash_window.mainloop()


def withdraw_cash():  # this function creates subsystem(withdraw cash)
    global withdraw_cash_window
    withdraw_cash_window = Tk()
    withdraw_cash_window.title('Withdraw Cash')
    withdraw_cash_window.configure(bg='Gray')
    withdraw_cash_window.geometry('500x400')

    # frames
    global frame_authenticate_withdraw
    frame_authenticate_withdraw = ttk.LabelFrame(withdraw_cash_window, text='Authenticate', width=450, height=180)
    frame_authenticate_withdraw.place(x=25, y=15)

    # labels, entries, and buttons
    lb_first_name = Label(frame_authenticate_withdraw, text='First Name: ')
    lb_first_name.place(x=5, y=10)

    global ent_first_name
    ent_first_name_var = StringVar()
    ent_first_name = Entry(frame_authenticate_withdraw, textvariable=ent_first_name_var, width=20)
    ent_first_name.place(x=150, y=10)

    lb_last_name = Label(frame_authenticate_withdraw, text='Last Name: ')
    lb_last_name.place(x=5, y=40)

    global ent_last_name
    ent_last_name_var = StringVar()
    ent_last_name = Entry(frame_authenticate_withdraw, textvariable=ent_last_name_var, width=20)
    ent_last_name.place(x=150, y=40)

    lb_acc_no = Label(frame_authenticate_withdraw, text='Acc Number: ')
    lb_acc_no.place(x=5, y=70)

    global ent_acc_no
    ent_acc_no_var = StringVar()
    ent_acc_no = Entry(frame_authenticate_withdraw, textvariable=ent_acc_no_var, width=20)
    ent_acc_no.place(x=150, y=70)

    # Button for authenticating the user's account.
    btn_authenticate_withdrawal = Button(frame_authenticate_withdraw, text='Authenticate', width=30, bg='Gray',
                                         command=lambda: authenticate_withdraw())
    btn_authenticate_withdrawal.place(x=45, y=110)

    withdraw_cash_window.mainloop()


def check_balance():  # this function creates subsystem(check balance)
    global check_balance_window, e_acc_no
    check_balance_window = Tk()
    check_balance_window.title('Check Balance')
    check_balance_window.configure(bg='Gray')
    check_balance_window.geometry('500x250')
    check_balance_window.resizable(width=False, height=False)

    # frames
    frame_check_balance = ttk.LabelFrame(check_balance_window, text='Check Balance', width=400, height=200)
    frame_check_balance.place(x=45, y=25)

    # label
    lbl_acc_no = Label(frame_check_balance, text='Account Number: ')
    lbl_acc_no.place(x=10, y=20)

    # entry
    e_acc_no_var = StringVar()
    e_acc_no = Entry(frame_check_balance, textvariable=e_acc_no_var, width=15)
    e_acc_no.place(x=180, y=20)

    # button
    button_check_balance = Button(frame_check_balance, text='Check Balance', width=20, height=3, bg='Gray',
                                  command=lambda: current_balance())
    button_check_balance.place(x=75, y=70)

    check_balance_window.mainloop()

# ====================================================


main_window = Tk()
# ======================= Settings ====================
main_window.title('Banking System(Simulation).')
main_window.geometry('800x350')
main_window.resizable(width=False, height=False)
main_window.configure(bg='Light Gray')
# =====================================================

# ======================= Title label =================
title_label = Label(main_window, text='WELCOME TO PYTHONISTA BANK OF KENYA.', font=('Tilda Petite', 15),
                    bg='Light Gray')
title_label.pack(side=TOP)
# =====================================================

# ======================= Frame to hold buttons =======
button_frame = ttk.LabelFrame(main_window, text='Bank\'s Agent Services.', width=600, height=210)
button_frame.place(x=100, y=70)
# =====================================================

# ======================= Buttons on button frame =====
btn_create_acc = Button(button_frame, text='Create Account', width=12, command=lambda: create_account())
btn_create_acc.place(x=26, y=35)

btn_deposit = Button(button_frame, text='Deposit', width=12, command=lambda: deposit_cash())
btn_deposit.place(x=216, y=35)

btn_withdraw = Button(button_frame, text='Withdraw', width=12, command=lambda: withdraw_cash())
btn_withdraw.place(x=406, y=35)

btn_check_balance = Button(button_frame, text='Check Balance', width=12, command=lambda: check_balance())
btn_check_balance.place(x=26, y=100)

btn_about = Button(button_frame, text='About', width=12, command=lambda: about_us())
btn_about.place(x=216, y=100)

btn_exit = Button(button_frame, text='EXIT', width=12, command=lambda: exit_program())
btn_exit.place(x=406, y=100)
# ======================================================

main_window.mainloop()
